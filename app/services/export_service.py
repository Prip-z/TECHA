from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.model import (
    Event,
    EventPlayer,
    Game,
    GameParticipant,
    Player,
    ShootingRound,
    StaffUser,
    Table,
    Testament,
    TestamentTarget,
    VotingNomination,
    VotingRound,
    VotingVote,
)
from app.schema.game import GameDraftExportRequest

STATUS_LABELS = {
    "preparation": "Подготовка",
    "voting": "Голосование",
    "revote": "Переголосование",
    "shooting": "Стрельба",
    "testament": "Завещание",
    "finished": "Завершено",
}

RESULT_LABELS = {
    "civilian_win": "Победа мирных",
    "mafia_win": "Победа мафии",
    "ppk_civilian_win": "ППК победа мирных",
    "ppk_mafia_win": "ППК победа мафии",
    "draw": "Ничья",
}

ROLE_LABELS = {
    "civilian": "Мирный",
    "mafia": "Мафия",
    "don": "Дон",
    "sheriff": "Шериф",
}


def _label(value: str | None, mapping: dict[str, str]) -> str | None:
    if value is None:
        return None
    return mapping.get(value, value)


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)


def _add_header_row(worksheet: Worksheet, values: list[str]) -> None:
    worksheet.append(values)
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "export"


def build_event_export(db: Session, event: Event) -> tuple[str, bytes]:
    workbook = Workbook()
    summary = workbook.active
    if not isinstance(summary, Worksheet):
        raise ValueError("Worksheet creation failed")
    summary.title = "Вечер"

    _add_header_row(summary, ["Поле", "Значение"])
    summary.append(["ID", event.id])
    summary.append(["Название", event.name])
    summary.append(["Дата", event.date.isoformat()])
    summary.append(["Тип", "Турнир" if event.type.value == "tournament" else "Обычный"])
    summary.append(["Цена за игру", event.price_per_game])

    registrations = (
        db.query(EventPlayer, Player)
        .join(Player, Player.id == EventPlayer.player_id)
        .filter(EventPlayer.event_id == event.id)
        .order_by(Player.nick.asc(), Player.id.asc())
        .all()
    )

    players_sheet = workbook.create_sheet("Игроки")
    _add_header_row(players_sheet, ["ID игрока", "Ник", "Имя", "Телефон", "Соцсеть", "Игр сыграно", "Оплачено"])
    for registration, player in registrations:
        players_sheet.append(
            [
                player.id,
                player.nick,
                player.name,
                player.phone,
                player.social_link,
                registration.games_played,
                registration.paid_amount,
            ]
        )
    if not registrations:
        players_sheet.append(["Нет данных", None, None, None, None, None, None])

    games = (
        db.query(Game, Table, StaffUser)
        .join(Table, Table.id == Game.table_id)
        .join(StaffUser, StaffUser.id == Game.host_staff_id)
        .filter(Game.event_id == event.id)
        .order_by(Game.game_number.asc(), Game.game_id.asc())
        .all()
    )

    games_sheet = workbook.create_sheet("Игры")
    _add_header_row(games_sheet, ["ID игры", "Номер игры", "Стол", "Ведущий", "Статус", "Результат", "Начало", "Окончание", "Протесты"])
    for game, table, host in games:
        games_sheet.append(
            [
                game.game_id,
                game.game_number,
                table.name,
                host.name,
                _label(game.status.value, STATUS_LABELS),
                _label(game.result.value if game.result else None, RESULT_LABELS),
                game.started_at.isoformat() if game.started_at else None,
                game.finished_at.isoformat() if game.finished_at else None,
                game.protests,
            ]
        )
    if not games:
        games_sheet.append(["Нет данных", None, None, None, None, None, None, None, None])

    for worksheet in workbook.worksheets:
        _autosize_columns(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = f"event_{event.id}_{_safe_filename(event.name)}.xlsx"
    return filename, buffer.getvalue()


def build_game_export(db: Session, game: Game, draft: GameDraftExportRequest | None = None) -> tuple[str, bytes]:
    workbook = Workbook()
    summary = workbook.active
    if not isinstance(summary, Worksheet):
        raise ValueError("Worksheet creation failed")
    summary.title = "Игра"

    event = db.get(Event, game.event_id)
    table = db.get(Table, game.table_id)
    host = db.get(StaffUser, game.host_staff_id)

    _add_header_row(summary, ["Поле", "Значение"])
    summary.append(["ID игры", game.game_id])
    summary.append(["Номер игры", game.game_number])
    summary.append(["ID вечера", game.event_id])
    summary.append(["Название вечера", event.name if event else None])
    summary.append(["Дата вечера", event.date.isoformat() if event else None])
    summary.append(["Стол", table.name if table else None])
    summary.append(["Ведущий", host.name if host else None])
    summary.append(["Статус", _label(game.status.value, STATUS_LABELS)])
    summary.append(["Результат", _label(game.result.value if game.result else None, RESULT_LABELS)])
    summary.append(["Начало", game.started_at.isoformat() if game.started_at else None])
    summary.append(["Окончание", game.finished_at.isoformat() if game.finished_at else None])
    summary.append(["Протесты", game.protests])

    participants = (
        db.query(GameParticipant, Player)
        .join(Player, Player.id == GameParticipant.player_id)
        .filter(GameParticipant.game_id == game.game_id)
        .order_by(GameParticipant.seat_number.asc(), GameParticipant.id.asc())
        .all()
    )
    seat_to_participant = {participant.seat_number: participant for participant, _ in participants}
    seat_to_player = {participant.seat_number: player for participant, player in participants}

    participants_sheet = workbook.create_sheet("Участники")
    _add_header_row(participants_sheet, ["Место", "ID игрока", "Ник", "Имя", "Фолы", "Баллы", "Доп. баллы", "Роль", "Жив"])
    for participant, player in participants:
        participants_sheet.append(
            [
                participant.seat_number,
                player.id,
                player.nick,
                player.name,
                participant.fouls,
                participant.score,
                participant.extra_score,
                _label(participant.role.value, ROLE_LABELS),
                "Да" if participant.is_alive else "Нет",
            ]
        )
    if not participants:
        participants_sheet.append(["Нет данных", None, None, None, None, None, None, None, None])

    voting_sheet = workbook.create_sheet("Голосование")
    _add_header_row(voting_sheet, ["Раунд", "Переголосование", "Ничья", "Поднятие", "Завершен", "Выбывший ID", "Номинации", "Голоса"])
    if draft and draft.votes:
        for voting_round in sorted(draft.votes, key=lambda item: item.round):
            nomination_text = ", ".join(
                f"{seat_to_player[seat].nick}({seat_to_player[seat].id})"
                for seat in voting_round.nominations
                if seat in seat_to_player
            )
            rendered_votes: list[str] = []
            for voter_seat, target_seat in sorted(voting_round.votes.items(), key=lambda item: int(item[0])):
                voter_player = seat_to_player.get(int(voter_seat))
                if voter_player is None:
                    continue
                if isinstance(target_seat, str) and target_seat.upper() == "X":
                    rendered_votes.append(f"{voter_player.nick}->X")
                    continue
                target_player = seat_to_player.get(int(target_seat))
                rendered_votes.append(f"{voter_player.nick}->{target_player.id if target_player else target_seat}")
            vote_text = ", ".join(rendered_votes)

            eliminated_player_id = None
            if not voting_round.isTie and not voting_round.isRevote and voting_round.votes:
                counts: dict[int, int] = {}
                for target_seat in voting_round.votes.values():
                    if isinstance(target_seat, str):
                        continue
                    counts[int(target_seat)] = counts.get(int(target_seat), 0) + 1
                if counts:
                    max_votes = max(counts.values())
                    leaders = [seat for seat, count in counts.items() if count == max_votes]
                    if len(leaders) == 1 and leaders[0] in seat_to_participant:
                        eliminated_player_id = seat_to_participant[leaders[0]].player_id

            voting_sheet.append(
                [
                    voting_round.round,
                    "Да" if voting_round.isRevote else "Нет",
                    "Да" if voting_round.isTie else "Нет",
                    "Да" if voting_round.liftApplied else "Нет",
                    "Да",
                    eliminated_player_id,
                    nomination_text,
                    vote_text,
                ]
            )
    else:
        voting_rounds = db.query(VotingRound).filter(VotingRound.game_id == game.game_id).order_by(VotingRound.round_number.asc()).all()
        if voting_rounds:
            for voting_round in voting_rounds:
                nominations = (
                    db.query(VotingNomination, Player)
                    .join(Player, Player.id == VotingNomination.player_id)
                    .filter(VotingNomination.voting_round_id == voting_round.id)
                    .order_by(Player.id.asc())
                    .all()
                )
                votes = (
                    db.query(VotingVote, Player)
                    .join(Player, Player.id == VotingVote.voter_player_id)
                    .filter(VotingVote.voting_round_id == voting_round.id)
                    .order_by(Player.id.asc())
                    .all()
                )
                nomination_text = ", ".join(f"{player.nick}({player.id})" for _, player in nominations)
                vote_text = ", ".join(f"{player.nick}->{vote.target_player_id if vote.target_player_id is not None else 'X'}" for vote, player in votes)
                voting_sheet.append(
                    [
                        voting_round.round_number,
                        "Да" if voting_round.is_revote else "Нет",
                        "Да" if voting_round.is_tie else "Нет",
                        "Да" if voting_round.is_lift_applied else "Нет",
                        "Да" if voting_round.is_completed else "Нет",
                        voting_round.eliminated_player_id,
                        nomination_text,
                        vote_text,
                    ]
                )
        else:
            voting_sheet.append(["Нет данных", None, None, None, None, None, None, None])

    shooting_sheet = workbook.create_sheet("Стрельба")
    _add_header_row(shooting_sheet, ["Раунд", "ID стрелка", "ID цели", "Промах"])
    draft_shots = [shot for shot in (draft.shots if draft else []) if shot.shooterSeat is not None or shot.targetSeat is not None]
    if draft_shots:
        for shooting_round in sorted(draft_shots, key=lambda item: item.round):
            shooter_player = seat_to_player.get(shooting_round.shooterSeat) if shooting_round.shooterSeat is not None else None
            target_player = None
            if isinstance(shooting_round.targetSeat, int):
                target_player = seat_to_player.get(shooting_round.targetSeat)
            shooting_sheet.append(
                [
                    shooting_round.round,
                    shooter_player.id if shooter_player else None,
                    target_player.id if target_player else None,
                    "Да" if shooting_round.targetSeat in (None, "X") else "Нет",
                ]
            )
    elif draft and draft.nights:
        for night in sorted(draft.nights, key=lambda item: item.round):
            target_player = seat_to_player.get(night.killedSeat) if night.killedSeat is not None else None
            shooting_sheet.append(
                [
                    night.round,
                    None,
                    target_player.id if target_player else None,
                    "Да" if night.killedSeat is None else "Нет",
                ]
            )
    else:
        shooting_rounds = (
            db.query(ShootingRound)
            .filter(ShootingRound.game_id == game.game_id)
            .order_by(ShootingRound.round_number.asc(), ShootingRound.id.asc())
            .all()
        )
        if shooting_rounds:
            for shooting_round in shooting_rounds:
                shooting_sheet.append(
                    [
                        shooting_round.round_number,
                        shooting_round.shooter_player_id,
                        shooting_round.target_player_id,
                        "Да" if shooting_round.is_miss else "Нет",
                    ]
                )
        else:
            shooting_sheet.append(["Нет данных", None, None, None])

    testament_sheet = workbook.create_sheet("Завещание")
    _add_header_row(testament_sheet, ["ID игрока", "Позиция цели", "ID цели"])
    if draft and draft.testament and draft.testament.sourceSeat is not None:
        source_player = seat_to_player.get(draft.testament.sourceSeat)
        if source_player:
            if not draft.testament.targetSeats:
                testament_sheet.append([source_player.id, None, None])
            else:
                for index, seat in enumerate(draft.testament.targetSeats, start=1):
                    target_player = seat_to_player.get(seat)
                    testament_sheet.append([source_player.id, index, target_player.id if target_player else None])
        else:
            testament_sheet.append(["Нет данных", None, None])
    else:
        testaments = db.query(Testament).filter(Testament.game_id == game.game_id).order_by(Testament.id.asc()).all()
        if testaments:
            for testament in testaments:
                targets = (
                    db.query(TestamentTarget)
                    .filter(TestamentTarget.testament_id == testament.id)
                    .order_by(TestamentTarget.position.asc(), TestamentTarget.id.asc())
                    .all()
                )
                if not targets:
                    testament_sheet.append([testament.player_id, None, None])
                    continue
                for target in targets:
                    testament_sheet.append([testament.player_id, target.position, target.target_player_id])
        else:
            testament_sheet.append(["Нет данных", None, None])

    for worksheet in workbook.worksheets:
        _autosize_columns(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = f"game_{game.game_id}_event_{game.event_id}.xlsx"
    return filename, buffer.getvalue()
